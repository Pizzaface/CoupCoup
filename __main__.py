import shutil
import asyncio
import configparser
import inspect
import json
from pathlib import Path
from typing import Type

import aiometer
import orjson
import pandas as pd
from async_timeout import timeout
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from loguru import logger
from pandas import DataFrame
from tqdm.asyncio import tqdm

from lib.constants import GLOBAL_COUPON_PROVIDERS
from stores.lib import BaseStore
from utils.spreadsheets import write_grouped_rows_with_colors, clean_workbook
import coupons
import stores
from offline_folium import offline   # noqa

from stores.lib.constants import HEADERS
from utils.config import get_config
from utils.geocoding import determine_store_paths
from utils.matching import match_multiple_columns

logger.remove()

# Log all errors, warnings, and info to the console
logger.add(tqdm.write, level='DEBUG')

Worksheet.to_list = lambda ws: list(ws.iter_rows(values_only=True))


async def main():
    Path('output/stores').mkdir(exist_ok=True, parents=True)
    section = _setup_config()

    await _handle_stores(section)
    await _handle_coupons(section)

    await _compare_products()
    await determine_store_paths()


async def _compare_products():
    wb = load_workbook('output/stores.xlsx')
    sheet_names = wb.sheetnames

    newspaper_coupons = get_global_coupons(sheet_names, wb)

    tasks = []
    # get the sales
    for sheet_name in sheet_names:
        if sheet_name.endswith(
            (*GLOBAL_COUPON_PROVIDERS, '-coupons', '-matchups')
        ):
            continue

        tasks.append(
            run_matchups_for_store(
                newspaper_coupons, sheet_name, sheet_names, wb
            )
        )

    results = await asyncio.gather(*tasks)

    for i, result in enumerate(results):
        if isinstance(result, Exception):
            logger.error(f'Error comparing products: {result}')
            continue

    clean_workbook(wb)

    wb.save('output/stores.xlsx')

    _split_sheets_by_store(wb)

    wb.close()


def _split_sheets_by_store(wb: Workbook):
    if Path('output/stores').exists():
        shutil.rmtree('output/stores')

    Path('output/stores').mkdir(exist_ok=True, parents=True)

    # split the stores into separate files
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name.endswith(('coupons', 'com')):
            continue

        # convert the sheet to a dataframe
        sales = wb[sheet_name]
        values = list(sales.values)[1:]
        sales = DataFrame(values, columns=HEADERS if not sheet_name.endswith('matchups') else [*HEADERS, "Matched Field", "Matched Value", 'Match Percentage', 'Matched Rows'])

        # save it as a csv
        sales.to_csv(f'output/stores/{sheet_name}.csv', index=False)


def get_global_coupons(sheet_names: list[str], wb: Workbook):
    global_coupons = []

    for sheet_name in sheet_names:
        if not sheet_name.endswith(GLOBAL_COUPON_PROVIDERS):
            continue

        sheet = wb[sheet_name]
        values = list(sheet.values)[1:]
        global_coupons.extend(values)

    global_coupons = DataFrame(global_coupons, columns=HEADERS)

    if global_coupons.empty:
        logger.error('No coupons found in the global coupon sheets')

    return global_coupons


async def run_matchups_for_store(
    newspaper_coupons: DataFrame,
    sheet_name: str,
    sheet_names: list[str],
    wb: Workbook,
):
    sales, total_coupons = _get_sales_and_coupons(
        newspaper_coupons, sheet_name, sheet_names, wb
    )

    try:
        # compare the products
        matches = match_multiple_columns(
            total_coupons,
            sales,
            ['brand_name', 'product_name', 'product_variety'],
            ['brand_name', 'product_name', 'product_variety'],
            threshold=90,
            limit=None,
        )
    except Exception as e:
        logger.error(f'Error comparing products for {sheet_name}: {e}')
        return

    sheet = wb.create_sheet(f'{sheet_name}-matchups', 0)
    rows = dataframe_to_rows(matches, index=False, header=True)

    write_grouped_rows_with_colors(rows, sheet)


def _get_sales_and_coupons(
    newspaper_coupons: DataFrame,
    sheet_name: str,
    sheet_names: list[str],
    wb: Workbook,
):
    if f'{sheet_name}-matchups' in sheet_names:
        del wb[f'{sheet_name}-matchups']

    sales = wb[sheet_name]
    values = list(sales.values)[1:]
    sales = DataFrame(values, columns=HEADERS)
    total_coupons = DataFrame()

    if f'{sheet_name}-coupons' in sheet_names:
        values = wb[f'{sheet_name}-coupons'].to_list()[1:]
        total_coupons = DataFrame(values, columns=HEADERS)

    if isinstance(newspaper_coupons, DataFrame):
        total_coupons = pd.concat(
            [total_coupons, newspaper_coupons],
            ignore_index=True,
            sort=False,
        )

    return sales, total_coupons


async def _handle_coupon_site(coupon, is_retry: bool = False):
    logger.info(f'Grabbing Coupons for Source: {coupon.__name__}')

    try:
        async with timeout(240) as cm:
            async with coupon(cm=cm) as coupon_obj:
                try:
                    await coupon_obj.scrape()
                except (asyncio.TimeoutError, asyncio.CancelledError) as e:
                    if is_retry:
                        logger.error(f'Timeout error scraping coupons: {e}')
                        return

                    logger.error(f'Timeout error scraping coupons: {e} - retrying...')
                    return await _handle_coupon_site(coupon, is_retry=True)

                except Exception as e:
                    logger.error(f'Error scraping coupons: {e}')
    except (asyncio.TimeoutError, asyncio.CancelledError) as e:
        if is_retry:
            logger.error(f'Timeout error scraping coupons: {e}')
            return

        logger.error(f'Timeout error scraping coupons: {e} - retrying...')
        return await _handle_coupon_site(coupon, is_retry=True)
    except Exception as e:
        logger.error(f'Error scraping coupons: {e}')


async def _handle_coupons(section: configparser.SectionProxy):
    included_coupons = orjson.loads(section.get('COUPON_SOURCES', '[]'))
    stores_at_once = section.getint('STORES_AT_ONCE', 2)

    if included_coupons:
        coupon_objs = inspect.getmembers(coupons, inspect.isclass)
        coupon_objs = [
            coupon
            for coupon_name, coupon in coupon_objs
            if coupon_name in included_coupons
            or coupon_name in GLOBAL_COUPON_PROVIDERS
            or coupon_name.removesuffix('Coupons') in included_coupons
        ]

        await aiometer.run_on_each(
            async_fn=_handle_coupon_site, args=coupon_objs, max_at_once=stores_at_once
        )

    else:
        logger.info(
            'No coupon sources included in config.ini - skipping coupon scraping.'
        )


def _setup_config():
    config = get_config()

    try:
        section = config['config']
    except KeyError:
        raise Exception(
            'No [config] section found in config.ini - please create one.'
        )

    if 'GOOGLE_PROJECT_ID' in section:
        import vertexai as genai

        genai.init(project=section['GOOGLE_PROJECT_ID'])
    elif 'GOOGLE_API_KEY' in section:
        import google.generativeai as genai

        genai.configure(api_key=section['GOOGLE_API_KEY'])
    else:
        raise Exception(
            "No Google API key found in config.ini - please add one under 'OPENAI_KEY' or 'GOOGLE_API_KEY' in the [config] section."
        )

    return section


async def _handle_stores(section):
    stores_at_once = section.getint('STORES_AT_ONCE', 2)
    included_stores = json.loads(section.get('INCLUDED_STORES', '[]'))

    if not included_stores:
        print('No stores included in config.ini - skipping store scraping.')
        return

    store_objs = [
        store
        for store_name, store in inspect.getmembers(stores, inspect.isclass)
        if store_name in included_stores
    ]

    await aiometer.run_on_each(
        async_fn=_run_store, args=store_objs, max_at_once=stores_at_once
    )

    print('Finished scraping stores')


async def _run_store(store: Type['BaseStore'], is_retry: bool = False):
    try:
        async with timeout(240) as cm:
            async with store(cm=cm) as store_obj:
                if store_obj.check_current_data():
                    logger.info(f'Store {store_obj._store_name} has up-to-date data (skipping).')
                    return

                logger.info(f'Grabbing Sales for Store: {store_obj._store_name}')
                try:
                    await store_obj.handle_flyers()
                except Exception as e:
                    logger.error(f'Error scraping store: {e}')
                    pass
    except (asyncio.TimeoutError, asyncio.CancelledError) as e:
        if is_retry:
            logger.error(f'Timeout error scraping store: {e}')
            return

        logger.error(f'Timeout error scraping store: {e} - retrying...')
        return await _run_store(store, is_retry=True)

    except Exception as e:
        logger.error(f'Error scraping store: {e}')

if __name__ == '__main__':
    loop = asyncio.new_event_loop()
    loop.run_until_complete(main())
