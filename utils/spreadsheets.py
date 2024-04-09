from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font

from utils.random_utils import random_hex_color_code


def write_grouped_rows_with_colors(rows, sheet):
    rows_to_colors = {}
    for r_idx, row in enumerate(rows, 1):
        if r_idx == 1:
            sheet.append(row)
            continue

        for c_idx, value in enumerate(row, 1):
            if isinstance(value, Cell):
                value = value.value

            cell = sheet.cell(row=r_idx, column=c_idx, value=value)

            matching_row_index = row[-1]
            if matching_row_index not in rows_to_colors:
                random_color_code, font_color = random_hex_color_code()
                while random_color_code in rows_to_colors.values():
                    random_color_code, font_color = random_hex_color_code()

                rows_to_colors[matching_row_index] = {
                    'fill': PatternFill('solid', fgColor=random_color_code),
                    'font': Font(color=font_color),
                }

            cell.fill = rows_to_colors[matching_row_index]['fill']
            cell.font = rows_to_colors[matching_row_index]['font']


def clean_workbook(wb):
    # remove empty sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if len(list(sheet.values)) <= 1:
            del wb[sheet_name]
