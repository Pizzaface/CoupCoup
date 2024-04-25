from __future__ import annotations

import asyncio
import json
import os
from typing import Dict, List
from zipfile import BadZipFile

import aiometer
import dateutil.parser
import orjson
import pandas as pd
from httpx import AsyncHTTPTransport
from loguru import logger
from more_itertools import chunked
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from pydantic import BaseModel
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm.asyncio import tqdm as tqdm_asyncio

from lib.RetryTransport import RetryTransport
from lib.constants import GLOBAL_COUPON_PROVIDERS
from stores.lib.constants import HEADERS, FILTER_KEYS
from utils.call_ai_model_gemini import extract_products_using_gemini
from utils.config import get_config


class Store(BaseModel):
    weekly_ad: List[Dict] | None = []
    headers: List[str] = HEADERS
    filter_keys: List[str] | None = FILTER_KEYS
    _store_name: str | None = None
    store_config: Dict[str, str] | None = None

    class Config:
        populate_by_name = True
        arbitrary_types_allowed = True
        extra = 'allow'

    processing_queue: list[dict] = []

    def __init__(self, cm, *args, **kwargs):
        super().__init__(**kwargs)

        self.timer_cm = cm
        config = get_config()

        store_name = self._store_name.lower()
        if store_name.endswith('-coupons'):
            store_name = store_name.removesuffix('-coupons')

        if store_name not in config.sections() and self._store_name.lower() not in GLOBAL_COUPON_PROVIDERS:
            raise Exception(
                f'{store_name} config not found - please check your config file'
            )
        elif store_name in config.sections():
            self.store_config = config[store_name]

        self.items_at_once = config['config'].getint('items_at_once', 5)

    async def __aenter__(self):
        self.pbar = tqdm_asyncio([], desc=self._store_name)
        self.pbar.disable = False
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        pass

    @property
    def httpx_transport(self):
        return RetryTransport(
            AsyncHTTPTransport(retries=3),
            max_attempts=10,
            retry_status_codes=[429, 500, 502, 503, 504],
        )

    def check_current_data(self):
        try:
            wb = load_workbook(self.excel_file_path)
        except (BadZipFile, KeyError, FileNotFoundError):
            return False

        if self._store_name not in wb.sheetnames:
            return False

        sheet = wb[self._store_name]
        if sheet.max_row <= 1:
            return False

        valid_to_col = next((cell.column for cell in sheet[1] if cell.value == 'Valid To'), None)
        if valid_to_col is None:
            return False

        for row in sheet.iter_rows(min_row=2):
            valid_to_cell = next((cell for cell in row if cell.column == valid_to_col and cell.value), None)
            if valid_to_cell:
                try:
                    dateutil.parser.parse(valid_to_cell.value)
                    return True
                except ValueError:
                    continue
        return False

    @property
    def logger(self):
        logger.remove()
        logger.add(
            tqdm_asyncio.write,
            colorize=True,
        )
        return logger.patch(lambda record: record.update(name=self._store_name))

    @property
    def excel_file_path(self) -> str:
        return 'output/stores.xlsx'

    def get_title_header(self) -> List[str]:
        return [h.replace('_', ' ').title() for h in self.headers]

    def reset_worksheet(self):
        if not os.path.exists(self.excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self._store_name
        else:
            try:
                wb = load_workbook(self.excel_file_path)
            except (BadZipFile, KeyError):
                wb = Workbook()
                ws = wb.active
                ws.title = self._store_name

            if self._store_name in wb.sheetnames:
                wb.remove(wb[self._store_name])

            ws = wb.create_sheet(title=self._store_name)

        ws.append(self.get_title_header())
        wb.save(self.excel_file_path)

    def add_row_to_store_worksheet(self, data: Dict):
        row = {
            col: data.get(col, 'N/A')
            if col not in ['requires_store_card']
            else bool(data.get(col))
            for col in self.headers
        }

        book = load_workbook(self.excel_file_path)
        sheet = book[self._store_name]

        df = pd.DataFrame([row])

        for row_obj in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row_obj)

        book.save(self.excel_file_path)

    def clean_worksheet(self):
        book = load_workbook(self.excel_file_path)
        sheet = book[self._store_name]
        dim_holder = DimensionHolder(worksheet=sheet)
        double = Side(border_style='thin', color='000000')

        for col in range(sheet.min_column, sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                sheet,
                min=col,
                max=col,
                bestFit=True,
                width=max([len(header) for header in self.headers]) + 5,
            )

        sheet.column_dimensions = dim_holder

        # format first row
        for cell in sheet[1]:
            cell.font = cell.font.copy(
                bold=True, color='FFFFFF', name='Segoe UI'
            )
            cell.border = cell.border.copy(
                left=double, right=double, top=double, bottom=double
            )
            cell.alignment = cell.alignment.copy(
                horizontal='center', vertical='center'
            )
            cell.fill = cell.fill.copy(fgColor='424242', patternType='solid')

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value is None:
                    cell.value = 'N/A'

                if cell.value != 'N/A':
                    try:
                        int(cell.value)
                    except ValueError:
                        try:
                            float(cell.value)
                        except ValueError:
                            if cell.value.lower() in ['true', 'false']:
                                cell.value = bool(cell.value)

                            cell.value = cell.value.encode(
                                'latin1', 'ignore'
                            ).decode('unicode_escape', 'ignore')
                        else:
                            cell.value = float(cell.value)
                            cell.style = 'Currency'
                            cell.number_format = '0.00'
                    else:
                        cell.value = int(cell.value)
                        cell.style = 'Comma'
                        cell.number_format = '#,##0'

                cell.alignment = cell.alignment.copy(
                    horizontal='center', vertical='center'
                )
        book.save(self.excel_file_path)

    async def process_queue(self, is_reprocess: bool = False):
        if not self.processing_queue:
            self.logger.info('No items to process')
            return

        if not is_reprocess:
            self.reset_worksheet()

        self.logger.info(f'Processing {len(self.processing_queue)} items')
        # Process the queue asynchronously and write to the Excel sheet
        reprocess_queue = []
        chunked_queue = [chunk for chunk in chunked(self.processing_queue, 2)]
        tasks = []

        for batch in chunked_queue:
            tasks.append(
                {
                    'prompt_jinja_template_path': 'get_individual_products.jinja',
                    'user_input': batch,
                    'logger': self.logger,
                }
            )

        self.pbar.reset(total=len(tasks))
        self.pbar.set_description(f'Processing {self._store_name}')
        self.pbar.refresh()

        self.timer_cm.shift(20 * len(tasks))
        async with aiometer.amap(async_fn=extract_products_using_gemini, args=tasks, max_at_once=self.items_at_once) as results:
            async for result_obj in results:
                if (
                    issubclass(result_obj.__class__, Exception)
                    or not isinstance(result_obj, tuple)
                    or not result_obj
                ):
                    continue

                self.pbar.update(1)
                self.timer_cm.shift(10)
                products, user_input = result_obj
                if isinstance(user_input, str):
                    try:
                        user_input = orjson.loads(user_input)
                    except json.JSONDecodeError:
                        self.logger.error(
                            f'Unable to decode user_input: {user_input}'
                        )
                        continue

                if not products:
                    self.logger.warning(
                        f'No products found for user input: {user_input} - adding to reprocess queue'
                    )
                    reprocess_queue.extend(user_input)
                    continue

                rows_to_add = []
                for i, product in enumerate(products):
                    if all(
                        [
                            product.get(key)
                            in ['N/A', None, 'COUPON', 'MANUFACTURER_COUPON']
                            for key in self.headers
                        ]
                    ):
                        self.logger.debug(
                            f'No valid data found for product: {product} - adding to reprocess queue'
                        )
                        reprocess_queue.extend(user_input)
                        continue

                    brand_names = []
                    if '|' in (product.get('brand_name') or ''):
                        brand_names = product['brand_name'].split('|')

                    if brand_names:
                        for brand_name in brand_names:
                            product['brand_name'] = brand_name.strip()
                            rows_to_add.append(product)
                    else:
                        rows_to_add.append(product)

                    try:
                        self.logger.debug(
                            f'Adding product to Excel: {product["brand_name"]} {product["product_name"]}'
                        )
                        if len(rows_to_add) >= 30:
                            for row in rows_to_add:
                                self.add_row_to_store_worksheet(row)

                            rows_to_add = []

                    except Exception as e:
                        self.logger.error(f'Error writing to Excel: {e}')
                        self.logger.error(
                            f'Adding product to reprocess queue: {product}'
                        )
                        reprocess_queue.extend(user_input)

                for row in rows_to_add:
                    self.add_row_to_store_worksheet(row)

        if reprocess_queue and not is_reprocess:
            self.processing_queue = reprocess_queue
            self.logger.info(f'Reprocessing {len(reprocess_queue)} items')
            self.timer_cm.shift(20 * len(reprocess_queue))

            return await self.process_queue(is_reprocess=True)
        elif reprocess_queue and is_reprocess:
            self.logger.error(
                f'Unable to process {len(reprocess_queue)} items after reprocessing. Please check logs.'
            )
            self.timer_cm.shift(5)
            await asyncio.sleep(2)

        self.logger.info('Finished processing queue')

        # now, we need to clean the worksheet
        self.clean_worksheet()


class CouponBaseStore(Store):
    ...
